from openpyxl import load_workbook


wb = load_workbook("sample.xlsx")
ws = wb.active

# 몇 점 이상
for row in ws.iter_rows(min_row=2):            
    if int(row[1].value) >60:
        print(row[0].value,"번 학생은 영어 천재")
        
# 영어 목록을 컴퓨터로 변환
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"




wb.save("sample_modified.xlsx")
        
        
        
        
