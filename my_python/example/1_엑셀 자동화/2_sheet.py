from openpyxl import Workbook


wb = Workbook()
ws = wb.create_sheet() #새로운 시트를 기본이름으로 생성
ws.title = "MySheet" #시트 이름변경
# ws.sheet.color = "ff66ff" #색상 변경 rgb 형태로 넣어야함

# sheet, mysheet, yoursheet
ws1 = wb.create_sheet("your_sheet") #주어진 이름으로 시트 생성
def new_func():
    return 2

ws2 = wb.create_sheet("NewSheet",new_func()) # type: ignore # 2번째 인덱스에 시트 생성

new_ws = wb["NewSheet"]#Dict 형태로 시트에 접근l

# print(wb.sheetnames) #모든 5시트 이름 확인

#시트 복사
new_ws["A1"] = "text" # type: ignore # newk_ws의 텍스토 A1에 입력
def new_func1():
    return new_ws

target = wb.copy_worksheet(new_func1()) # type: ignore #
target.title = "Copied Sheet"


wb.save("sample.xlsx")
wb.close()



