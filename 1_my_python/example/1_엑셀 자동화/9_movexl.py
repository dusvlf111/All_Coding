from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

'''
# 위치로 이동 시킴
ws.move_range("B1:C11", rows=0, cols=1) 


#국어 입력​⁡
ws["B1"].value = "국어" 

'''


ws.move_range('C1:C11',rows=5,cols=-1)






wb.save("sample_korean.xlsx")