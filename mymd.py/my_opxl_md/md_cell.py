from random import randint
from openpyxl import Workbook


wb = Workbook()
ws = wb.active



#반복문을 이용해서 랜덤 숫자 채우기
class random_cell:
    index = 1
    for x in range(1,11):
        for y in range(1,11):
            ws.cell(column=x, row=y, value=randint(0,100)) # 0부터 100 사이의 숫자
            ws.cell(row=x, column=y, value=index)
            index += 1
            
def mad_ap():
    ws.append(["번호","영어","수학"])