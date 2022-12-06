from openpyxl import Workbook
import datetime
wb = Workbook()
ws = wb.active



ws["A1"] = datetime.datetime.today() # 오늘의 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)"          # 합계
ws["A3"] = "=AVERAGE(1, 2, 3)"      # 평균


ws["A4"] = 10       
ws['A5'] = 20
ws['A6'] = '=SUM(A4:A5)' #30만들기



wb.save('sample_formula.xlsx')
