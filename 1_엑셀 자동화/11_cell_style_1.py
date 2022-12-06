from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

wb = load_workbook("sample.xlsx")
ws = wb.active


a1 = ws["A1"] #번호
b1 = ws["B1"] #영어
c1 = ws["C1"] #수학



# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5


# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50


# 스타일 적용
#글자색은 빨간 이텔릭 두껍게
a1.font = Font(color="FF0000", italic=True, bold=True)
#                       글자색  폰트모양    취소선
b1.font = Font(color="CC33FF", name="Arial", strike=True)
#                       글자색    사이즈      밑줄
c1.font = Font(color = "0000FF", size=20, underline="single")

# 태두리 적용



thin_border = Border(
    left=Side(style="thin")
    ,right=Side(style="thin")
    ,top=Side(style="thin")
    ,bottom=Side(style="thin")
    )


a1.border = thin_border
b1.border = thin_border
c1.border = thin_border




wb.save('sample_style.xlsx')

