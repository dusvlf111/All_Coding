from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook("sample.xlsx")
ws = wb.active


a1 = ws["A1"]
b1 = ws["B2"]
c1 = ws["C1"]



# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5


# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50


# 스타일 적용
#글자색은 빨간 이텔릭 두껍게
a1.font = Font(color="FF0000", italic=True, bold=True)

wb.save('sample_style.xlsx')

