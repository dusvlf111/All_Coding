from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을 땐 none 을 출력

# row = 1,2,3,4 ....
#column = A(1),B(2),C(3),D(4) ....
ws.cell(row=1, column=1)

print(ws.cell(column=1, row=1).value) #print(ws["A1"])
print(ws.cell(column=2, row=1).value)#print(ws["B1"])

#ws.cell(column=3, row=1, value=10) # ws["c1"] = 10 과 동일
c = ws.cell(column=3, row=1, value=10) # ws["c1"] = 10 과 동일
print(c.value) #ws["C1"].value

from random import randrange


#반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1,11):
    for y in range(1,11):
        #ws.cell(column=x, row=y, value=randint(0,100)) # 0부터 100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1
        
        

wb.save("sample.xlsx")