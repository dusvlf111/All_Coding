from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active


ws.merge_cells("B2:D2") #B2 부터 D2까지 합치겠음
ws['B2'].value = "merged Cell"


wb.save("sample_merge.xlsx")