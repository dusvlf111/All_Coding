from openpyxl import load_workbook

from openpyxl.styles  import (
    Font, Border, Side, PatternFill, Alignment
    )

wb = load_workbook("sample.xlsx")
ws = wb.active


a1 = ws["A1"] #번호
b1 = ws["B1"] #영어
c1 = ws["C1"] #수학



# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5


# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50


# 스타일 적용
#글자색은 빨간 이텔릭 두껍게
a1.font = Font(color="FF0000", italic=True, bold=True)
#                       글자색  폰트모양    취소선
b1.font = Font(color="CC33FF", name="Arial", strike=True)
#                       글자색    사이즈      밑줄
c1.font = Font(color = "0000FF", size=20, underline="single")

# 태두리 적용

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
    )


a1.border = thin_border
b1.border = thin_border
c1.border = thin_border
# 태두리 적용

# 90점이 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 cell 에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        #                          horizontal = center left top bottom
        if cell.column == 1: # A번호열은 제외
            continue
        # cell 이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            #           배경색을 바꿈       색          채우는 형태
            cell.fill = PatternFill(fgColor = "00FF00",fill_type="solid")
            #           폰트    색상 변경
            cell.font = Font(color="FF0000")
            
            
# 틀 고정 이동해도 고정됨 따라온다
ws.freeze_panes = "B2" #B2 기준으로 틀 고정






        

wb.save('sample_style.xlsx')

